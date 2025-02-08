
import openpyxl

workbook = openpyxl.load_workbook('C:\\Users\\shiva\\PycharmProjects\\GfGpythonLearning\\Selenium\\Excel_Runner.xlsx')
active_sheet = workbook.active
current_cell  = active_sheet.cell(row=2, column = 3)
print(current_cell.value)

list_of_data = []
my_dict_excel = {}
# active_sheet.cell(row=5, column = 1).value = "Test 4"
# current_cell  = active_sheet.cell(row=5, column = 1)
# print(current_cell.value)
#
# active_sheet.cell(row=5, column = 2).value = "Sierra"
# current_cell  = active_sheet.cell(row=5, column = 2)
# print(current_cell.value)

row_count=active_sheet.max_row
column_count = active_sheet.max_column

for rows in range(1,row_count+1):
    for columns in range(1,column_count+1):
        # current_cell = active_sheet.cell(row=rows, column=columns)
        # print(current_cell.value, end=' ')
        my_dict_excel[active_sheet.cell(row=1, column=columns).value] = active_sheet.cell(row=rows, column=columns).value

    #print("\n")
    list_of_data.append(my_dict_excel)

print(list_of_data)




