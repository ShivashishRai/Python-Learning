
#package to import excel files
import openpyxl

#locating the excel file location into an path object
path = '../ExcelPython.xlsx'
workbook_location = openpyxl.load_workbook(path)
myactivesheet = workbook_location.active

total_row = myactivesheet.max_row
total_column = myactivesheet.max_column

# print(total_column)
# print(total_row)
# cell_value = myactivesheet.cell(row = 5, column= 4)
# print(cell_value.value)

#Iterating over each element in the excel sheet
# for i in range(2, total_row+1):
#     for j in range (1,total_column+1):
#         cell_value = myactivesheet.cell(row=i,column=j)
#         #print(cell_value.value)
#         print(cell_value.value[0])


#read mutiple cells from cell name
cell_value = myactivesheet['A1':'D10']
for cell1, cell2,cell3,cell4 in cell_value:
    print(cell1.value,cell2.value,cell3.value,cell4.value)

